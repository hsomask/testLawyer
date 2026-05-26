"""
Excel 导出：计算明细 + 审计信息（openpyxl）。

表头顺序：费用类目、计算基数、利率标准、起始日、截止日、天数、金额。
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

from legal_calc.common.lpr_json_file import default_lpr_1y_json_path
from legal_calc.private_lending import PrivateLendingRequest
from legal_calc.rental.models import RentalRequest
from legal_calc.report_models import CalculationResult, ReportLineItem

# 文档列名（严格顺序）
REPORT_HEADERS = (
    "费用类目",
    "计算基数",
    "利率标准",
    "起始日",
    "截止日",
    "天数",
    "金额",
)

# 律政蓝系淡底（打印友好，近似 Tailwind blue-100）
_HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="1E3A5F")
_MONEY_ALIGN = Alignment(horizontal="right", vertical="center")
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _fee_cell_text(ln: ReportLineItem) -> str:
    """费用类目列：保留阶段说明于同一单元格，满足单列头时的可读性。"""
    if not ln.stage_description.strip():
        return ln.fee_category
    return f"{ln.fee_category}｜{ln.stage_description}"


def _input_snapshot_to_str(input_snapshot: BaseModel | dict[str, Any] | str | None) -> str:
    if input_snapshot is None:
        return ""
    if isinstance(input_snapshot, str):
        return input_snapshot
    if isinstance(input_snapshot, BaseModel):
        return json.dumps(
            input_snapshot.model_dump(mode="json"),
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    return json.dumps(input_snapshot, ensure_ascii=False, indent=2, default=str)


def _resolve_lpr_raw(
    lpr_raw_snapshot: str | None,
    lpr_file_path: Path | None,
) -> tuple[str, str]:
    """
    返回 (来源说明, 原始文本)。
    """
    if lpr_raw_snapshot is not None:
        return ("调用方提供的 LPR 快照", lpr_raw_snapshot)
    path = lpr_file_path if lpr_file_path is not None else default_lpr_1y_json_path()
    if not path.is_file():
        return ("LPR 文件", f"[文件不存在] {path}")
    text = path.read_text(encoding="utf-8")
    return (str(path.resolve()), text)


def _style_detail_sheet(ws) -> None:
    """表头加粗、淡蓝底；金额与基数列右对齐两位小数。"""
    max_col = len(REPORT_HEADERS)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    money_fmt = "#,##0.00"
    for row in range(2, ws.max_row + 1):
        # B=计算基数, G=金额
        ws.cell(row=row, column=2).alignment = _MONEY_ALIGN
        ws.cell(row=row, column=2).number_format = money_fmt
        ws.cell(row=row, column=7).alignment = _MONEY_ALIGN
        ws.cell(row=row, column=7).number_format = money_fmt
        ws.cell(row=row, column=6).alignment = _CENTER_ALIGN

    ws.freeze_panes = "A2"
    widths = {"A": 42, "B": 14, "C": 28, "D": 12, "E": 12, "F": 8, "G": 14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def _style_audit_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            c = ws.cell(row=row, column=col)
            if row == 1:
                c.font = _HEADER_FONT
                c.fill = _HEADER_FILL
            if col == 2:
                c.alignment = _WRAP_ALIGN


def _fill_detail(ws, lines: list[ReportLineItem]) -> None:
    ws.append(list(REPORT_HEADERS))
    for ln in lines:
        ws.append(
            [
                _fee_cell_text(ln),
                ln.principal_base,
                ln.rate_standard,
                ln.period_start.isoformat(),
                ln.period_end.isoformat(),
                ln.day_count,
                ln.amount,
            ]
        )


def _append_private_lending_detail_summary(ws, result: CalculationResult) -> None:
    """PRD §3.1：明细表底部追加利息小计、冲抵后本金、本息合计。"""
    if result.interest_subtotal is None:
        return
    ws.append([])
    ws.append(
        [
            "【利息小计】",
            "",
            "",
            "",
            "",
            "",
            result.interest_subtotal,
        ]
    )
    ws.append(
        [
            "【冲抵后剩余本金】",
            "",
            "",
            "",
            "",
            "",
            result.remaining_principal,
        ]
    )
    ws.append(
        [
            "【本息合计】",
            "",
            "",
            "",
            "",
            "",
            result.total_principal_and_interest,
        ]
    )


def _append_rental_detail_summary(ws, result: CalculationResult) -> None:
    """房屋租赁：明细表底部追加汇总区（欠租本金、各滞纳金、占用费、最终总计）。"""
    rs = result.rental_summary
    if rs is None:
        return
    ws.append([])
    summary_rows = [
        ("【应收租金小计】", rs.rent_receivable_subtotal),
        ("【已支付租金合计】", rs.paid_rent_amount),
        ("【欠租本金小计】", rs.arrears_principal_subtotal),
        ("【租金滞纳金小计】", rs.rent_late_fee_subtotal),
        ("【水电费滞纳金小计】", rs.utility_late_fee_subtotal),
        ("【物业费滞纳金小计】", rs.property_late_fee_subtotal),
        ("【其他费用滞纳金小计】", rs.other_late_fee_subtotal),
        ("【房屋占用费小计】", rs.occupancy_fee_subtotal),
        ("【最终总计】", rs.grand_total),
    ]
    for label, value in summary_rows:
        ws.append([label, "", "", "", "", "", value])


def _fill_audit(
    ws,
    *,
    result: CalculationResult,
    input_snapshot: BaseModel | dict[str, Any] | str | None,
    lpr_raw_snapshot: str | None,
    lpr_file_path: Path | None,
) -> None:
    snap = _input_snapshot_to_str(input_snapshot)
    src, lpr_text = _resolve_lpr_raw(lpr_raw_snapshot, lpr_file_path)

    ws.append(["字段", "内容"])
    ws.append(["RULE_VERSION", result.rule_version])
    ws.append(["Input_Snapshot (JSON)", snap or "（未提供）"])
    ws.append(["LPR_Data_Source", src])
    ws.append(["LPR_Raw_JSON", lpr_text])
    ws.append([])
    ws.append(["assumptions_used", "\n".join(result.assumptions_used)])
    ws.append(["messages", "\n".join(result.messages)])
    ws.append(["line_amount_sum", str(result.line_amount_sum())])
    if result.interest_subtotal is not None:
        ws.append(["interest_subtotal", str(result.interest_subtotal)])
        ws.append(["remaining_principal", str(result.remaining_principal)])
        ws.append(["total_principal_and_interest", str(result.total_principal_and_interest)])
    if result.rental_summary is not None:
        ws.append([])
        ws.append(
            [
                "rental_grand_total",
                str(result.rental_summary.grand_total),
            ]
        )
        ws.append(
            [
                "rental_summary",
                result.rental_summary.model_dump_json(indent=2),
            ]
        )


def write_report_workbook(
    result: CalculationResult,
    path: Path | None = None,
    *,
    input_snapshot: BaseModel | dict[str, Any] | str | None = None,
    lpr_raw_snapshot: str | None = None,
    lpr_file_path: Path | None = None,
    include_private_lending_totals: bool = False,
    include_rental_totals: bool = False,
) -> Path | BytesIO:
    """
    生成 xlsx：Sheet「计算明细」+「审计信息」。

    - 审计页含 RULE_VERSION、Input_Snapshot、计算所用 LPR 原始 JSON（默认读包内 ``lpr_1y_cny.json``）。
    - 金额、基数：保留两位小数、居右。

    若 ``path`` 为 None，返回 ``BytesIO``。
    """
    wb = Workbook()
    ws_detail = wb.active
    assert ws_detail is not None
    ws_detail.title = "计算明细"
    _fill_detail(ws_detail, result.lines)
    if include_private_lending_totals:
        _append_private_lending_detail_summary(ws_detail, result)
    if include_rental_totals:
        _append_rental_detail_summary(ws_detail, result)
    _style_detail_sheet(ws_detail)

    ws_audit = wb.create_sheet("审计信息", 1)
    _fill_audit(
        ws_audit,
        result=result,
        input_snapshot=input_snapshot,
        lpr_raw_snapshot=lpr_raw_snapshot,
        lpr_file_path=lpr_file_path,
    )
    _style_audit_sheet(ws_audit)

    if path is not None:
        wb.save(path)
        return path
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# 兼容旧名称
write_workbook = write_report_workbook


def export_rental_workbook(req: RentalRequest, result: CalculationResult) -> BytesIO:
    """房屋租赁计算书（Excel + 审计页，含请求快照、LPR JSON、汇总区、rental_summary）。"""
    out = write_report_workbook(
        result, None, input_snapshot=req, include_private_lending_totals=False, include_rental_totals=True,
    )
    assert isinstance(out, BytesIO)
    return out


def export_private_lending_workbook(req: PrivateLendingRequest, result: CalculationResult) -> BytesIO:
    """
    将 ``calculate_private_lending`` 的结果导出为标准 Excel 计算书（内存 BytesIO）。

    审计页自动附带请求快照与默认 LPR JSON。
    """
    out = write_report_workbook(result, None, input_snapshot=req, include_private_lending_totals=True)
    assert isinstance(out, BytesIO)
    return out
