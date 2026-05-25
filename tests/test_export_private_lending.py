"""民间借贷 Excel 导出：审计页完整性 + 金额一致性测试。"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl

from legal_calc.export import export_private_lending_workbook
from legal_calc.private_lending import PrivateLendingRequest, calculate_private_lending


def _make_req() -> PrivateLendingRequest:
    return PrivateLendingRequest(
        principal=Decimal("10000"),
        loan_date=date(2020, 9, 1),
        end_date=date(2020, 10, 31),
        agreed_annual_rate=Decimal("0.12"),
        filing_date=date(2020, 10, 15),
    )


def test_private_lending_audit_sheet_has_all_required_fields() -> None:
    """审计信息页包含 PRD 要求的全部字段。"""
    req = _make_req()
    result = calculate_private_lending(req)
    bio = export_private_lending_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]
    labels = [audit.cell(row=r, column=1).value for r in range(2, audit.max_row + 1)]
    required = [
        "RULE_VERSION",
        "Input_Snapshot (JSON)",
        "LPR_Data_Source",
        "LPR_Raw_JSON",
        "assumptions_used",
        "messages",
        "line_amount_sum",
        "interest_subtotal",
        "remaining_principal",
        "total_principal_and_interest",
    ]
    for field in required:
        assert field in labels, f"审计信息缺少字段: {field}"


def test_private_lending_line_amount_sum_matches_detail_lines() -> None:
    """审计信息中的 line_amount_sum 应等于明细行金额加总。"""
    req = _make_req()
    result = calculate_private_lending(req)
    bio = export_private_lending_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]

    sum_from_audit = None
    for row in range(2, audit.max_row + 1):
        if audit.cell(row=row, column=1).value == "line_amount_sum":
            sum_from_audit = audit.cell(row=row, column=2).value
            break

    assert sum_from_audit is not None, "审计信息中缺少 line_amount_sum"
    expected = str(result.line_amount_sum())
    assert sum_from_audit == expected, f"line_amount_sum 不一致: {sum_from_audit} != {expected}"


def test_private_lending_interest_subtotal_matches_interest_lines() -> None:
    """审计信息中的 interest_subtotal 应等于利息行金额之和。"""
    req = _make_req()
    result = calculate_private_lending(req)
    bio = export_private_lending_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]

    audit_interest_subtotal = None
    for row in range(2, audit.max_row + 1):
        if audit.cell(row=row, column=1).value == "interest_subtotal":
            audit_interest_subtotal = audit.cell(row=row, column=2).value
            break

    assert audit_interest_subtotal is not None, "审计信息中缺少 interest_subtotal"
    expected = str(result.interest_subtotal)
    assert audit_interest_subtotal == expected


def test_private_lending_total_equals_interest_plus_remaining() -> None:
    """total_principal_and_interest == interest_subtotal + remaining_principal。"""
    req = _make_req()
    result = calculate_private_lending(req)
    assert result.total_principal_and_interest is not None
    assert result.interest_subtotal is not None
    assert result.remaining_principal is not None
    expected_total = result.interest_subtotal + result.remaining_principal
    assert result.total_principal_and_interest == expected_total, (
        f"本息合计不一致: {result.total_principal_and_interest} != {expected_total}"
    )
