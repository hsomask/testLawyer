"""房屋租赁 Excel 导出：审计页完整性 + 不含借贷合计行 + 金额一致性测试。"""

from datetime import date
from decimal import Decimal

import openpyxl

from legal_calc.export import REPORT_HEADERS, export_rental_workbook
from legal_calc.rental import RentalRequest, calculate_rental

_PRIVATE_LENDING_SUMMARY_LABELS = [
    "【利息小计】",
    "【冲抵后剩余本金】",
    "【本息合计】",
]


def _make_req() -> RentalRequest:
    return RentalRequest(
        monthly_rent=Decimal("3000.00"),
        arrears_period_start=date(2025, 1, 1),
        arrears_period_end=date(2025, 1, 31),
        rent_due_day_of_month=26,
        contract_termination_date=date(2025, 3, 1),
        actual_vacate_date=None,
        filing_date=date(2025, 4, 1),
    )


def test_rental_excel_has_detail_and_audit_sheets() -> None:
    """房屋租赁 Excel 应包含「计算明细」和「审计信息」两个 Sheet。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    assert "计算明细" in wb.sheetnames, "缺少「计算明细」Sheet"
    assert "审计信息" in wb.sheetnames, "缺少「审计信息」Sheet"


def test_rental_excel_has_detail_headers() -> None:
    """房屋租赁 Excel 计算明细页存在表头。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    ws = wb["计算明细"]
    headers = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    assert headers == REPORT_HEADERS, f"表头不一致: {headers}"


def test_rental_excel_no_private_lending_summary_rows() -> None:
    """房屋租赁 Excel 不应出现民间借贷专属的三行合计。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    ws = wb["计算明细"]
    col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    for label in _PRIVATE_LENDING_SUMMARY_LABELS:
        assert label not in col_a, f"租赁 Excel 不应出现民间借贷合计行: {label}"


def test_rental_audit_sheet_has_all_required_fields() -> None:
    """房屋租赁审计信息页包含 PRD 要求的全部字段（不含借贷专属三字段）。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]
    labels = [audit.cell(row=r, column=1).value for r in range(2, audit.max_row + 1)]

    required = [
        "RULE_VERSION",
        "Input_Snapshot (JSON)",
        "LPR_Data_Source",
        "LPR_Raw_JSON",
        "assumptions_used",
        "messages",
        "line_amount_sum",
    ]
    for field in required:
        assert field in labels, f"审计信息缺少字段: {field}"

    prohibited = ["interest_subtotal", "remaining_principal", "total_principal_and_interest"]
    for field in prohibited:
        assert field not in labels, f"租赁审计信息不应包含借贷专属字段: {field}"


def test_rental_line_amount_sum_matches_detail_lines() -> None:
    """审计信息中的 line_amount_sum 应等于明细行金额加总。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]

    sum_from_audit = None
    for row in range(2, audit.max_row + 1):
        if audit.cell(row=row, column=1).value == "line_amount_sum":
            sum_from_audit = audit.cell(row=row, column=2).value
            break

    assert sum_from_audit is not None, "审计信息中缺少 line_amount_sum"
    expected = str(result.line_amount_sum())
    assert sum_from_audit == expected, f"line_amount_sum 不一致: {sum_from_audit} != {expected}"


def test_rental_excel_has_messages_in_audit() -> None:
    """房屋租赁审计信息中 messages 字段有值（租赁通常有 warnings）。"""
    req = _make_req()
    result = calculate_rental(req)
    assert len(result.messages) > 0, "租赁计算应有 messages"
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]

    messages_val = None
    for row in range(2, audit.max_row + 1):
        if audit.cell(row=row, column=1).value == "messages":
            messages_val = audit.cell(row=row, column=2).value
            break

    assert messages_val is not None, "审计信息中缺少 messages"
    assert len(messages_val) > 0, "messages 不应为空字符串"


def test_rental_detail_lines_distinguishable_by_fee_category() -> None:
    """房屋租赁明细行可按 fee_category 区分不同费用类目。"""
    from legal_calc.rental.models import RentalExtraFeeItem

    req = RentalRequest(
        monthly_rent=Decimal("3000.00"),
        arrears_period_start=date(2025, 1, 1),
        arrears_period_end=date(2025, 1, 31),
        rent_due_day_of_month=26,
        contract_termination_date=date(2025, 3, 1),
        actual_vacate_date=date(2025, 3, 15),
        filing_date=date(2025, 4, 1),
        extra_fee_items=[
            RentalExtraFeeItem(
                category="property", name="物业费", amount=Decimal("200"), due_date=date(2025, 1, 26)
            ),
            RentalExtraFeeItem(
                category="utility", name="水电费", amount=Decimal("100"), due_date=date(2025, 1, 26)
            ),
        ],
    )
    result = calculate_rental(req)
    categories = {ln.fee_category for ln in result.lines}
    assert "租金滞纳金" in categories, f"应包含租金滞纳金，实际类目: {categories}"
    assert "房屋占用费" in categories, f"应包含房屋占用费，实际类目: {categories}"
    assert "物业费滞纳金" in categories, f"应包含物业费滞纳金，实际类目: {categories}"
    assert "水电费滞纳金" in categories, f"应包含水电费滞纳金，实际类目: {categories}"


_RENTAL_SUMMARY_LABELS = [
    "【应收租金小计】",
    "【已支付租金合计】",
    "【欠租本金小计】",
    "【租金滞纳金小计】",
    "【水电费滞纳金小计】",
    "【物业费滞纳金小计】",
    "【其他费用滞纳金小计】",
    "【房屋占用费小计】",
    "【最终总计】",
]


def test_rental_excel_detail_has_summary_rows() -> None:
    """房屋租赁 Excel 计算明细底部存在所有汇总行。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    ws = wb["计算明细"]
    col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    for label in _RENTAL_SUMMARY_LABELS:
        assert label in col_a, f"租赁 Excel 明细缺少汇总行: {label}"


def test_rental_excel_audit_has_rental_summary() -> None:
    """房屋租赁审计信息页包含 rental_summary JSON。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]
    labels = [audit.cell(row=r, column=1).value for r in range(2, audit.max_row + 1)]
    assert "rental_summary" in labels, "审计信息缺少 rental_summary"


def test_rental_excel_grand_total_matches_rental_summary() -> None:
    """Excel 中的【最终总计】等于 rental_summary.grand_total。"""
    req = _make_req()
    result = calculate_rental(req)
    assert result.rental_summary is not None
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    ws = wb["计算明细"]
    found = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "【最终总计】":
            found = ws.cell(row=row, column=7).value
            break
    assert found is not None, "未找到【最终总计】行"
    assert Decimal(str(found)) == result.rental_summary.grand_total, (
        f"Excel 最终总计不一致: {found} != {result.rental_summary.grand_total}"
    )


def test_rental_excel_audit_has_rental_grand_total() -> None:
    """房屋租赁审计信息页包含 rental_grand_total 字段。"""
    req = _make_req()
    result = calculate_rental(req)
    bio = export_rental_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    audit = wb["审计信息"]
    labels = [audit.cell(row=r, column=1).value for r in range(2, audit.max_row + 1)]
    assert "rental_grand_total" in labels, "审计信息缺少 rental_grand_total"


def test_rental_excel_grand_total_differs_from_line_amount_sum_when_paid() -> None:
    """有 paid_rent_amount 时，grand_total 与 line_amount_sum 不一致（明细行不含扣减）。"""
    req = RentalRequest(
        monthly_rent=Decimal("3000.00"),
        arrears_period_start=date(2025, 1, 1),
        arrears_period_end=date(2025, 1, 31),
        rent_due_day_of_month=26,
        contract_termination_date=date(2025, 3, 1),
        actual_vacate_date=date(2025, 3, 15),
        filing_date=date(2025, 4, 1),
        paid_rent_amount=Decimal("1000.00"),
    )
    result = calculate_rental(req)
    assert result.rental_summary is not None
    line_sum = result.line_amount_sum()
    grand = result.rental_summary.grand_total
    # line_amount_sum > grand_total（因为 line_amount_sum 含完整的应收租金）
    assert line_sum > grand, f"line_amount_sum={line_sum} 应大于 grand_total={grand}"
