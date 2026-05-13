from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl

from legal_calc.export import REPORT_HEADERS, export_private_lending_workbook, write_report_workbook
from legal_calc.private_lending import PrivateLendingRequest, calculate_private_lending
from legal_calc.report_models import CalculationResult, ReportLineItem


def test_write_report_workbook_creates_file() -> None:
    result = CalculationResult(
        ok=True,
        rule_version="test",
        assumptions_used=["a"],
        lines=[
            ReportLineItem(
                fee_category="测试",
                stage_description="行1",
                principal_base=Decimal("1.00"),
                rate_standard="N/A",
                period_start=date(2024, 1, 1),
                period_end=date(2024, 1, 2),
                day_count=2,
                amount=Decimal("0.50"),
            )
        ],
        messages=["m"],
    )
    req = PrivateLendingRequest(
        principal=Decimal("10000"),
        loan_date=date(2024, 1, 1),
        end_date=date(2024, 6, 1),
        agreed_annual_rate=Decimal("0.06"),
    )
    with TemporaryDirectory() as td:
        p = Path(td) / "out.xlsx"
        write_report_workbook(result, p, input_snapshot=req)
        assert p.exists()
        assert p.stat().st_size > 0

        wb = openpyxl.load_workbook(p)
        assert "计算明细" in wb.sheetnames
        assert "审计信息" in wb.sheetnames

        ws = wb["计算明细"]
        assert tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] == REPORT_HEADERS

        audit = wb["审计信息"]
        labels = [audit.cell(row=r, column=1).value for r in range(2, 6)]
        assert "RULE_VERSION" in labels
        assert "Input_Snapshot (JSON)" in labels
        assert "LPR_Data_Source" in labels
        assert "LPR_Raw_JSON" in labels


def test_private_lending_excel_includes_prd_3_1_summary_rows() -> None:
    req = PrivateLendingRequest(
        principal=Decimal("10000"),
        loan_date=date(2020, 9, 1),
        end_date=date(2020, 10, 31),
        repayments=[],
        agreed_annual_rate=Decimal("0.12"),
        filing_date=date(2020, 10, 15),
    )
    result = calculate_private_lending(req)
    bio = export_private_lending_workbook(req, result)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    ws = wb["计算明细"]
    col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "【利息小计】" in col_a
    assert "【冲抵后剩余本金】" in col_a
    assert "【本息合计】" in col_a

    audit = wb["审计信息"]
    keys = [audit.cell(row=r, column=1).value for r in range(2, audit.max_row + 1)]
    assert "interest_subtotal" in keys
    assert "remaining_principal" in keys
    assert "total_principal_and_interest" in keys


def test_bytes_io_export() -> None:
    result = CalculationResult(ok=True, rule_version="v", lines=[], messages=[])
    bio = write_report_workbook(result, None)
    bio.seek(0)
    wb = openpyxl.load_workbook(bio)
    assert "计算明细" in wb.sheetnames
